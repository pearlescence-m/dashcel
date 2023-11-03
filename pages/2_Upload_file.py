import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(
    page_title="dashcel",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'mailto:meruyert.alimaganbetova@gmail.com',
        'Report a bug': "mailto:meruyert.alimaganbetova@gmail.com",
        'About': "# dashcel project is licensed under the Apache License, Version 2.0 (c) 2023"
    }
)

st.markdown("# 📊 dashcel")
st.markdown("---")

uploaded_file = st.file_uploader("Choose an Excel file to convert: ")

sheets = []

if uploaded_file is not None:
    wb = load_workbook(uploaded_file.name, read_only=True)
    sheets = wb.sheetnames   
    options = st.multiselect(
        'Select Sheets from file:',
        sheets
        )
