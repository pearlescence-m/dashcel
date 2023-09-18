import sys
from PySide6 import QtCore, QtWidgets, QtGui
from openpyxl import load_workbook

class ExcelFileImporter(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Excel File Importer")
        self.setGeometry(100, 100, 600, 300)

        self.initUI()

    def initUI(self):
        # Create QLineEdit to display the selected file's path
        self.file_path_input = QtWidgets.QLineEdit(self)
        self.file_path_input.setGeometry(10, 10, 400, 30)
        self.file_path_input.setPlaceholderText("Selected File: No file selected")
        self.file_path_input.setReadOnly(True)

        # Create QPushButton for file selection
        self.select_button = QtWidgets.QPushButton("Select File", self)
        self.select_button.setGeometry(420, 10, 100, 30)
        self.select_button.clicked.connect(self.select_file)

        # # Create QCheckBox for selecting the first worksheet by default
        self.list_view_text = QtWidgets.QLabel("Select worksheets to convert into dashboards:", self)
        self.list_view_text.setGeometry(10, 50, 300, 30) 

        # Create QListView for selecting multiple worksheets
        self.worksheet_list_view = QtWidgets.QListView(self)
        self.worksheet_list_view.setGeometry(10, 90, 300, 80)
        self.worksheet_list_view.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)

        # Create QStandardItemModel
        self.worksheet_list_model = QtGui.QStandardItemModel(self.worksheet_list_view)
        self.worksheet_list_view.setModel(self.worksheet_list_model)

        # Create QPushButton for uploading file
        self.upload_button = QtWidgets.QPushButton("Upload File", self)
        self.upload_button.setGeometry(10, 180, 100, 30)
        self.upload_button.clicked.connect(self.upload_file)

    def select_file(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.ReadOnly  # Ensure the file is opened in read-only mode
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Excel File", "", 
                                                             "Excel Files (*.xlsx);;All Files (*)", options=options)

        if file_path:
            self.file_path_input.setText(f"{file_path}")
            self.select_button.setText("Change File")
            self.select_button.clicked.connect(self.clear_selection)

            # Clear existing worksheet selections
            self.worksheet_list_model.clear()

            # Load worksheet names into the list view
            wb = load_workbook(file_path, read_only=True)
            worksheet_names = wb.sheetnames
            for name in worksheet_names:
                item = QtGui.QStandardItem(name)
                item.setCheckable(True)
                item.setCheckState(QtCore.Qt.Checked)  # Initially checked
                self.worksheet_list_model.appendRow(item)

    def clear_selection(self):
        self.file_path_input.clear()
        self.select_button.setText("Select File")
        self.select_button.clicked.connect(self.select_file)
        # Clear all selections in the list view
        for row in range(self.worksheet_list_model.rowCount()):
            item = self.worksheet_list_model.item(row)
            if item is not None:
                item.setCheckState(QtCore.Qt.Unchecked)

    def upload_file(self):
        print("uploaded")
        pass

if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    window = ExcelFileImporter()
    window.show()
    sys.exit(app.exec())