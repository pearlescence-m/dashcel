from PySide6 import QtWidgets, QtGui, QtCore
from openpyxl import load_workbook

from menubar import LeftPaneWidget

class ExcelFileImporter(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel File Importer")
        self.setGeometry(100, 100, 600, 300)
        self.initUI()


    def initUI(self):
        # Create a central widget
        central_widget = QtWidgets.QWidget(self)
        self.setCentralWidget(central_widget)
        splitter = QtWidgets.QHBoxLayout(central_widget)

        # Create a LeftPaneWidget instance and add it to the left side of the splitter
        left_pane = LeftPaneWidget()
        splitter.addWidget(left_pane)

        # Create main content widget
        main_content = QtWidgets.QWidget()
        splitter.addWidget(main_content)

        # Set the central widget as the splitter
        central_layout = QtWidgets.QVBoxLayout(main_content)

        # Create a QHBoxLayout for "Select File" QPushButton and QLineEdit
        file_selection_layout = QtWidgets.QHBoxLayout()

        # Create QLineEdit to display the selected file's path
        self.file_path_input = QtWidgets.QLineEdit(self)
        self.file_path_input.setPlaceholderText("Selected File: No file selected")
        self.file_path_input.setReadOnly(True)

        # Create QPushButton for file selection
        self.select_button = QtWidgets.QPushButton("Select File", self)
        self.select_button.clicked.connect(self.select_file)
        self.select_button.setFixedSize(100, 25) 

        # Add the QPushButton and QLineEdit to the file_selection_layout
        file_selection_layout.addWidget(self.select_button)
        file_selection_layout.addWidget(self.file_path_input)

        # # Create QCheckBox for text label
        self.list_view_text = QtWidgets.QLabel("Select worksheets to convert into dashboards:", self)

        # Create QListView for selecting multiple worksheets
        self.worksheet_list_view = QtWidgets.QListView(self)
        self.worksheet_list_view.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)

        # Create QStandardItemModel
        self.worksheet_list_model = QtGui.QStandardItemModel(self.worksheet_list_view)
        self.worksheet_list_view.setModel(self.worksheet_list_model)

        # Create QPushButton for uploading file
        self.upload_button = QtWidgets.QPushButton("Upload File", self)
        self.upload_button.clicked.connect(self.upload_file)
        self.upload_button.setFixedSize(100, 25) 

        # Create the "Select/Deselect All" checkbox
        self.select_deselect_checkbox = QtWidgets.QCheckBox("Select All/None", self)
        self.select_deselect_checkbox.setChecked(True)
        self.select_deselect_checkbox.stateChanged.connect(self.toggle_select_all)

        # Flag to keep track of the state
        self.select_all_flag = True

        # Add widgets to the layout
        central_layout.addWidget(self.file_path_input)
        central_layout.addWidget(self.select_button)
        central_layout.addWidget(self.list_view_text)
        central_layout.addWidget(self.select_deselect_checkbox)
        central_layout.addWidget(self.worksheet_list_view)
        central_layout.addWidget(self.upload_button)

    
    def toggle_select_all(self, state):
        # Slot to toggle selecting/deselecting all items in the list view
        self.select_all_flag = not self.select_all_flag
        check_state = QtCore.Qt.Checked if self.select_all_flag else QtCore.Qt.Unchecked
        for row in range(self.worksheet_list_model.rowCount()):
            item = self.worksheet_list_model.item(row)
            if item is not None:
                item.setCheckState(check_state)


    def select_file(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.ReadOnly  # Ensure the file is opened in read-only mode
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Excel File", "",
                                                             "Excel Files (*.xlsx);;All Files (*)", options=options)

        if file_path:
            self.file_path_input.setText(f"{file_path}")
            self.select_button.setText("Change File")
            self.select_button.clicked.connect(self.clear_selection)

            # Clear existing worksheet selections
            self.worksheet_list_model.clear()

            # Load worksheet names into the list view
            wb = load_workbook(file_path, read_only=True)
            worksheet_names = wb.sheetnames
            for name in worksheet_names:
                item = QtGui.QStandardItem(name)
                item.setCheckable(True)
                item.setCheckState(QtCore.Qt.Checked)  # Initially checked
                self.worksheet_list_model.appendRow(item)


    def clear_selection(self):
        self.file_path_input.clear()
        self.select_button.setText("Select File")
        self.select_button.clicked.connect(self.select_file)
        # Clear all selections in the list view
        for row in range(self.worksheet_list_model.rowCount()):
            item = self.worksheet_list_model.item(row)
            if item is not None:
                item.setCheckState(QtCore.Qt.Unchecked)


    def upload_file(self):
        print("uploaded")
        pass